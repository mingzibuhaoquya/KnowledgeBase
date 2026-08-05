import csv
import re
import shutil
import uuid
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from fastapi import HTTPException, UploadFile

from app.core.config import settings
from app.services.ocr import ocr_service


@dataclass
class ParsedImage:
    filename: str
    file_path: str
    ocr_text: str = ""


@dataclass
class ParsedDocument:
    stored_filename: str
    file_path: str
    file_type: str
    file_hash: str
    file_size: int
    content_text: str
    images: list[ParsedImage] = field(default_factory=list)


class DocumentParser:
    allowed_suffixes = {".txt", ".md", ".docx", ".pdf", ".xlsx", ".csv"}

    async def save_and_parse(self, upload: UploadFile) -> ParsedDocument:
        suffix = Path(upload.filename or "document").suffix.lower()
        if suffix not in self.allowed_suffixes:
            raise HTTPException(status_code=400, detail="Unsupported file type. Please upload txt, md, docx, pdf, xlsx, or csv.")

        doc_dir = settings.upload_root / uuid.uuid4().hex
        image_dir = doc_dir / "images"
        image_dir.mkdir(parents=True, exist_ok=True)

        stored_filename = f"source{suffix or '.bin'}"
        target_path = doc_dir / stored_filename
        with target_path.open("wb") as buffer:
            shutil.copyfileobj(upload.file, buffer)
        file_size = target_path.stat().st_size
        if file_size == 0:
            shutil.rmtree(doc_dir, ignore_errors=True)
            raise HTTPException(status_code=400, detail="Uploaded file is empty.")
        file_hash = self._sha256(target_path)

        try:
            text = self._extract_text(target_path, suffix)
            images = self._extract_images(target_path, suffix, image_dir)
        except Exception as exc:
            text = f"Parse failed: {exc}"
            images = []
        ocr_lines: list[str] = []
        for image in images:
            image.ocr_text = ocr_service.extract_text(Path(image.file_path))
            if image.ocr_text:
                ocr_lines.append(image.ocr_text)

        if ocr_lines:
            text = f"{text}\n\nImage OCR:\n" + "\n".join(ocr_lines)

        return ParsedDocument(
            stored_filename=stored_filename,
            file_path=str(target_path),
            file_type=suffix.removeprefix(".") or "unknown",
            file_hash=file_hash,
            file_size=file_size,
            content_text=text.strip(),
            images=images,
        )

    def reparse_existing(self, file_path: str, file_type: str) -> str:
        path = Path(file_path)
        suffix = f".{file_type.lower().lstrip('.')}"
        if not path.exists():
            return ""
        if suffix not in self.allowed_suffixes:
            return ""
        return self._extract_text(path, suffix).strip()

    def _sha256(self, file_path: Path) -> str:
        import hashlib

        digest = hashlib.sha256()
        with file_path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    def _extract_text(self, file_path: Path, suffix: str) -> str:
        if suffix in {".txt", ".md"}:
            return file_path.read_text(encoding="utf-8", errors="ignore")
        if suffix == ".docx":
            return self._extract_docx_text(file_path)
        if suffix == ".pdf":
            return self._extract_pdf_text(file_path)
        if suffix == ".xlsx":
            return self._extract_xlsx_text(file_path)
        if suffix == ".csv":
            return self._extract_csv_text(file_path)
        return ""

    def _extract_docx_text(self, file_path: Path) -> str:
        try:
            from docx import Document
        except ImportError:
            return "[python-docx is not installed. Cannot parse docx text.]"

        doc = Document(str(file_path))
        sections: list[str] = []
        context_title = ""
        context_section = ""
        context_meta: dict[str, str] = {}
        pending_meta_key = ""
        table_index = 0
        for block in self._iter_docx_blocks(doc):
            if hasattr(block, "text"):
                text = block.text.strip()
                if not text:
                    continue
                sections.append(text)
                if pending_meta_key:
                    context_meta[pending_meta_key] = text
                    pending_meta_key = ""
                    continue
                if self._is_interface_heading(text):
                    context_title = text
                    context_meta = {}
                    context_section = ""
                if text == "交易方向":
                    pending_meta_key = "Transaction direction"
                    continue
                if text == "接口类型":
                    pending_meta_key = "Interface type"
                    continue
                if "请求报文" in text or "请求参数" in text or text.upper() == "REQUEST":
                    context_section = "REQUEST"
                elif "应答报文" in text or "响应报文" in text or "返回报文" in text or text.upper() == "RESPONSE":
                    context_section = "RESPONSE"
            else:
                table_index += 1
                table_text = self._format_docx_table(block, table_index, context_title, context_section, context_meta)
                if table_text:
                    sections.append(table_text)
        return "\n\n".join(sections)

    def _iter_docx_blocks(self, doc):
        from docx.table import Table
        from docx.text.paragraph import Paragraph
        from docx.oxml.table import CT_Tbl
        from docx.oxml.text.paragraph import CT_P

        body = doc.element.body
        for child in body.iterchildren():
            if isinstance(child, CT_P):
                yield Paragraph(child, doc)
            elif isinstance(child, CT_Tbl):
                yield Table(child, doc)

    def _format_docx_table(
        self,
        table,
        index: int,
        context_title: str = "",
        context_section: str = "",
        context_meta: dict[str, str] | None = None,
    ) -> str:
        rows: list[list[str]] = []
        for row in table.rows:
            cells = [self._clean_cell(cell.text) for cell in row.cells]
            cells = self._dedupe_repeated_cells(cells)
            if any(cells):
                rows.append(cells)
        if not rows:
            return ""

        title = self._table_title(rows, index)
        interface_name = self._interface_name(rows, context_title)
        lines = [f"[TABLE:{index}] {title}"]
        if interface_name:
            lines.append(f"Interface name: {interface_name}")
        for key, value in (context_meta or {}).items():
            if value:
                lines.append(f"{key}: {value}")
        if context_section:
            lines.append(f"Interface section: {context_section}")
        elif self._looks_like_request(rows):
            lines.append("Interface section: REQUEST")
        elif self._looks_like_response(rows):
            lines.append("Interface section: RESPONSE")
        header_index = self._header_index(rows)
        if header_index is not None:
            header = rows[header_index]
            lines.append("Columns: " + " | ".join(cell for cell in header if cell))
            for row_number, row in enumerate(rows[header_index + 1 :], start=1):
                values = self._normalize_row(row, len(header))
                if not any(values):
                    continue
                mapped = self._map_table_row(header, values)
                if mapped:
                    lines.append(f"Row {row_number}: " + " | ".join(f"{key}: {value}" for key, value in mapped))
                else:
                    lines.append(f"Row {row_number}: " + " | ".join(value for value in values if value))
        else:
            for row_number, row in enumerate(rows, start=1):
                values = [value for value in row if value]
                if values:
                    lines.append(f"Row {row_number}: " + " | ".join(values))
        lines.append(f"[/TABLE:{index}]")
        return "\n".join(lines)

    def _clean_cell(self, text: str) -> str:
        return re.sub(r"\s+", " ", text or "").strip()

    def _dedupe_repeated_cells(self, cells: list[str]) -> list[str]:
        deduped: list[str] = []
        previous = object()
        for cell in cells:
            if cell == previous:
                continue
            deduped.append(cell)
            previous = cell
        return deduped

    def _table_title(self, rows: list[list[str]], index: int) -> str:
        flat = " ".join(cell for row in rows[:3] for cell in row if cell)
        if "REQUEST" in flat.upper():
            return "Interface request fields"
        if "RESPONSE" in flat.upper():
            return "Interface response fields"
        if any(name in flat for name in ["中文域名", "标签名", "是否必输", "字段"]):
            return "Interface fields"
        return f"Table {index}"

    def _interface_name(self, rows: list[list[str]], context_title: str) -> str:
        flat = " ".join(cell for row in rows[:6] for cell in row if cell)
        patterns = [
            r"接口名称[:：]\s*([^|，,。；;\n]+)",
            r"接口名[:：]\s*([^|，,。；;\n]+)",
            r"接口[:：]\s*([^|，,。；;\n]+)",
            r"([A-Za-z0-9_./-]+(?:接口|Service|API))",
        ]
        for pattern in patterns:
            match = re.search(pattern, flat)
            if match:
                return match.group(1).strip()
        if context_title and ("接口" in context_title or "查询" in context_title or "申请" in context_title):
            return context_title[:120]
        return ""

    def _looks_like_request(self, rows: list[list[str]]) -> bool:
        flat = " ".join(cell for row in rows[:5] for cell in row if cell).upper()
        return "<REQUEST>" in flat or "REQUEST" in flat or "请求" in flat

    def _looks_like_response(self, rows: list[list[str]]) -> bool:
        flat = " ".join(cell for row in rows[:5] for cell in row if cell).upper()
        return "<RESPONSE>" in flat or "RESPONSE" in flat or "返回" in flat or "响应" in flat

    def _last_interface_title(self, paragraphs: list[str]) -> str:
        for text in reversed(paragraphs[-30:]):
            clean = text.strip()
            if not clean:
                continue
            if self._is_interface_heading(clean):
                return clean
        return ""

    def _is_interface_heading(self, text: str) -> bool:
        clean = text.strip()
        generic_labels = {
            "接口类型",
            "请求报文",
            "应答报文",
            "响应报文",
            "返回报文",
            "交易方向",
            "校验规则",
            "业务功能描述",
            "界面描述",
            "Webservice接口",
        }
        if clean in generic_labels:
            return False
        return len(clean) <= 120 and (
            "服务接口" in clean
            or "接口服务" in clean
            or "接口【" in clean
            or "接口[" in clean
            or "查询" in clean
            or "申请" in clean
            or re.search(r"^\d+(?:\.\d+)+", clean) is not None
        )

    def _header_index(self, rows: list[list[str]]) -> int | None:
        for index, row in enumerate(rows[:8]):
            joined = "|".join(row)
            if ("中文域名" in joined and "标签名" in joined) or ("字段" in joined and "说明" in joined):
                return index
            if sum(1 for cell in row if cell) >= 3 and any("说明" in cell or "备注" in cell for cell in row):
                return index
        return None

    def _normalize_row(self, row: list[str], size: int) -> list[str]:
        values = list(row[:size])
        if len(values) < size:
            values.extend([""] * (size - len(values)))
        return values

    def _map_table_row(self, header: list[str], row: list[str]) -> list[tuple[str, str]]:
        aliases = {
            "序号": "index",
            "中文域名": "field_cn",
            "中文名称": "field_cn",
            "标签名": "field_code",
            "字段名": "field_code",
            "长度": "length",
            "是否必输": "required",
            "说明": "description",
            "备注": "description",
        }
        mapped: list[tuple[str, str]] = []
        for name, value in zip(header, row):
            if not value:
                continue
            key = aliases.get(name.strip(), name.strip() or "value")
            mapped.append((key, value))
        return mapped

    def _extract_pdf_text(self, file_path: Path) -> str:
        try:
            from pypdf import PdfReader
        except ImportError:
            return "[pypdf is not installed. Cannot parse pdf text.]"

        reader = PdfReader(str(file_path))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    def _extract_xlsx_text(self, file_path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return "[openpyxl is not installed. Cannot parse xlsx text.]"

        workbook = load_workbook(str(file_path), data_only=True, read_only=True)
        sections: list[str] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            sections.append(f"Sheet: {sheet.title}")
            header = self._clean_row(rows[0])
            if header:
                sections.append("Columns: " + " | ".join(header))
            for row_index, row in enumerate(rows[1:] if header else rows, start=2 if header else 1):
                cells = self._clean_row(row)
                if not cells:
                    continue
                if header and len(header) == len(cells):
                    pairs = [f"{name}: {value}" for name, value in zip(header, cells)]
                    sections.append(f"Row {row_index}: " + " | ".join(pairs))
                else:
                    sections.append(f"Row {row_index}: " + " | ".join(cells))
        workbook.close()
        return "\n".join(sections)

    def _extract_csv_text(self, file_path: Path) -> str:
        text = file_path.read_text(encoding="utf-8-sig", errors="ignore")
        rows = list(csv.reader(text.splitlines()))
        if not rows:
            return ""
        sections = ["CSV Table"]
        header = self._clean_row(rows[0])
        if header:
            sections.append("Columns: " + " | ".join(header))
        for row_index, row in enumerate(rows[1:] if header else rows, start=2 if header else 1):
            cells = self._clean_row(row)
            if not cells:
                continue
            if header and len(header) == len(cells):
                sections.append(f"Row {row_index}: " + " | ".join(f"{name}: {value}" for name, value in zip(header, cells)))
            else:
                sections.append(f"Row {row_index}: " + " | ".join(cells))
        return "\n".join(sections)

    def _clean_row(self, row) -> list[str]:
        cells: list[str] = []
        for value in row:
            if value is None:
                continue
            text = str(value).strip()
            if text:
                cells.append(text)
        return cells

    def _extract_images(self, file_path: Path, suffix: str, image_dir: Path) -> list[ParsedImage]:
        if suffix != ".docx":
            return []

        images: list[ParsedImage] = []
        with zipfile.ZipFile(file_path) as archive:
            for item in archive.namelist():
                if not item.startswith("word/media/"):
                    continue
                source_name = Path(item).name
                target_name = f"{len(images) + 1:03d}_{source_name}"
                target_path = image_dir / target_name
                with archive.open(item) as source, target_path.open("wb") as target:
                    shutil.copyfileobj(source, target)
                images.append(ParsedImage(filename=target_name, file_path=str(target_path)))
        return images


document_parser = DocumentParser()
